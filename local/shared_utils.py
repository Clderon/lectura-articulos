#!/usr/bin/env python3
"""
Utilidades compartidas - Funciones reutilizadas de procesar_articulos.py
"""
import sys
import os
import re
import time
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI

# Cargar .env
load_dotenv()

DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
DEEPSEEK_BASE_URL = "https://api.deepseek.com"
DEEPSEEK_MODEL = "deepseek-chat"
TEMPERATURE = 0.1


# Patrones de secciones
SECTION_HEADER_PATTERNS = [
    r"abstract", r"resumen",
    r"keywords?|palabras\s+clave",
    r"introduction|introducción",
    r"background|antecedentes",
    r"methodology|metodología|method|método",
    r"results?|resultados?",
    r"discussion|discusión",
    r"conclusion|conclusión|conclusions|conclusiones",
]


def send_to_deepseek(client: OpenAI, messages: list, content: str, retries: int = 3) -> str:
    """Envía contenido a DeepSeek y retorna respuesta."""
    messages.append({"role": "user", "content": content})
    last_error = None

    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=DEEPSEEK_MODEL,
                messages=messages,
                temperature=TEMPERATURE,
            )
            reply = response.choices[0].message.content
            messages.append({"role": "assistant", "content": reply})
            return reply
        except Exception as e:
            last_error = e
            print(f"    [DeepSeek] Intento {attempt + 1}/{retries} falló: {e}")
            if attempt < retries - 1:
                wait = 5 * (attempt + 1)
                print(f"    Esperando {wait}s...")
                time.sleep(wait)

    raise RuntimeError(f"DeepSeek falló tras {retries} intentos: {last_error}")


def cut_at_references(text: str) -> str:
    """Corta el texto ANTES de la sección de referencias (busca genérica)."""
    lines = text.split('\n')

    # Palabras clave que indican inicio de referencias/bibliografía
    reference_keywords = ['references', 'bibliography', 'referencias', 'bibliografía', 'works cited', 'literatura citada']

    for i, line in enumerate(lines):
        line_lower = line.lower().strip()
        # Buscar si la línea contiene alguna palabra clave de referencias
        for keyword in reference_keywords:
            if keyword in line_lower:
                # Retornar todo ANTES de esta línea
                return '\n'.join(lines[:i])

    return text


def split_into_sections(text: str) -> dict:
    """Divide texto en secciones detectando encabezados."""
    lines = text.split("\n")
    sections = {}
    current_section = "Inicio"
    current_content = []

    for line in lines:
        line_lower = line.lower().strip()

        # Detectar encabezado de sección
        is_header = False
        for pattern in SECTION_HEADER_PATTERNS:
            if re.match(pattern, line_lower):
                # Guardar sección anterior
                if current_content:
                    content = "\n".join(current_content).strip()
                    if content:
                        sections[current_section] = content

                current_section = line.strip()
                current_content = []
                is_header = True
                break

        if not is_header:
            if line.strip():
                current_content.append(line)

    # Guardar última sección
    if current_content:
        content = "\n".join(current_content).strip()
        if content:
            sections[current_section] = content

    return sections if sections else {"Contenido": text}


def write_article_to_excel(excel_path: str, art_num: int, tabla1: dict, tabla2: dict):
    """Escribe datos en Excel usando openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError:
        print("[ERROR] openpyxl no está instalado")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws1 = wb["Revisión de Artículos"]
        ws2 = wb["Lectura Selectiva"]

        # art_num ahora es directamente el número de fila (no requiere +3)
        row = art_num

        ws1[f"B{row}"].value = tabla1.get("Artículo (Título)", "")
        ws1[f"C{row}"].value = tabla1.get("Autor(es)", "")
        ws1[f"D{row}"].value = tabla1.get("Año", "")
        ws1[f"E{row}"].value = tabla1.get("Pertinencia", 0)
        ws1[f"F{row}"].value = tabla1.get("Actualidad", 0)
        ws1[f"G{row}"].value = tabla1.get("Rigor", 0)
        ws1[f"H{row}"].value = tabla1.get("Fuente (Calidad)", 0)
        ws1[f"I{row}"].value = tabla1.get("Impacto", 0)
        ws1[f"J{row}"].value = tabla1.get("Ética", 0)
        ws1[f"L{row}"].value = tabla1.get("Veredicto", "")
        ws1[f"M{row}"].value = tabla1.get("Justificación", "")

        ws2[f"C{row}"].value = tabla2.get("Objetivo", "")
        ws2[f"D{row}"].value = tabla2.get("Método", "")
        ws2[f"E{row}"].value = tabla2.get("Hallazgo", "")
        ws2[f"F{row}"].value = tabla2.get("Limitación", "")
        ws2[f"G{row}"].value = tabla2.get("Referencia textual / Sección", "")
        ws2[f"H{row}"].value = tabla2.get("Paradigma IaC cubierto", "")
        ws2[f"I{row}"].value = tabla2.get("Herramienta IaC analizada", "")
        ws2[f"J{row}"].value = tabla2.get("Tipo de estudio", "")
        ws2[f"K{row}"].value = tabla2.get("Variables definidas", "")
        ws2[f"L{row}"].value = tabla2.get("Dimensiones", "")
        ws2[f"M{row}"].value = tabla2.get("Indicadores medibles", "")
        ws2[f"N{row}"].value = tabla2.get("Trabajo futuro declarado", "")
        ws2[f"O{row}"].value = tabla2.get("¿Generalizable a híbrido?", "")
        ws2[f"P{row}"].value = tabla2.get("Uso en tu tesis", "")

        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            ws2[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical='top')

        ws2.row_dimensions[row].height = 60

        wb.save(excel_path)
        wb.close()

        print(f"    OK Excel guardado (N°{art_num})")

    except Exception as e:
        print(f"    [ERROR] No se pudo escribir Excel: {e}")


def load_analyzed_articles(excel_path: str) -> dict:
    """Lee artículos ya analizados del Excel (respeta fórmulas)."""
    try:
        import openpyxl
    except ImportError:
        return {}

    try:
        # data_only=True para leer valores calculados, no fórmulas
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Leer de Revisión de Artículos: Columna B (Título) tiene valores reales
        if "Revisión de Artículos" in wb.sheetnames:
            ws = wb["Revisión de Artículos"]
            analyzed = {}

            # Leer solo columna B (Artículo/Título) que es valor real, no fórmula
            for row_idx in range(4, 44):  # Filas 4-43
                title_cell = ws[f"B{row_idx}"].value

                if title_cell:
                    try:
                        title_normalized = str(title_cell).lower().strip()
                        # Guardar con el índice de fila como N°
                        row_num = row_idx - 3  # Convertir fila a N°
                        analyzed[title_normalized] = (row_num, str(title_cell))
                    except (ValueError, TypeError):
                        pass

            wb.close()
            return analyzed

        wb.close()
        return {}

    except Exception as e:
        print(f"[WARNING] No se pudo leer Excel: {e}")
        return {}


def create_deepseek_client() -> OpenAI:
    """Crea cliente de DeepSeek."""
    return OpenAI(api_key=DEEPSEEK_API_KEY, base_url=DEEPSEEK_BASE_URL)
