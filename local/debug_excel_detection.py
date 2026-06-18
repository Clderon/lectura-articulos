#!/usr/bin/env python3
"""
Debug: Detecta dónde debería escribir en Excel sin llamar a DeepSeek
"""
import os
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

load_dotenv()

excel_path = os.getenv("EXCEL_PATH")

print("=" * 70)
print("DEBUG - Detección de filas en Excel")
print("=" * 70)

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Revisión de Artículos"]

    print(f"\nArchivo Excel: {excel_path}\n")

    # 1. Detectar dinámicamente dónde está el encabezado
    print("[PASO 1] Buscando encabezado...")
    header_row = None
    for row_idx in range(1, 10):
        cell_a = ws[f"A{row_idx}"].value
        cell_b = ws[f"B{row_idx}"].value

        print(f"  Fila {row_idx}: A={cell_a}, B={cell_b}")

        if cell_a and isinstance(cell_a, str) and any(x in str(cell_a).lower() for x in ["n°", "numero"]):
            header_row = row_idx
            print(f"  ✓ ENCABEZADO ENCONTRADO en fila {row_idx} (columna A)")
            break
        elif cell_b and isinstance(cell_b, str) and any(x in str(cell_b).lower() for x in ["autor", "artículo", "título"]):
            header_row = row_idx
            print(f"  ✓ ENCABEZADO ENCONTRADO en fila {row_idx} (columna B)")
            break

    if header_row is None:
        print("  ✗ No encontró encabezado, asumiendo fila 2")
        content_start = 2
    else:
        content_start = header_row + 1
        print(f"\n  Contenido empieza en fila: {content_start}\n")

    # 2. Mostrar estado de filas
    print("[PASO 2] Analizando filas desde fila", content_start)
    print()

    for row_idx in range(content_start, content_start + 10):
        num_en_a = ws[f"A{row_idx}"].value
        titulo_en_b = ws[f"B{row_idx}"].value

        # SOLO verificar columna B (Título)
        estado = "LLENA" if titulo_en_b else "VACÍA"
        print(f"  Fila {row_idx}: A={num_en_a}, B={titulo_en_b}, Estado={estado}")

        # Si es la primera vacía (B vacío)
        if num_en_a is not None and not titulo_en_b:
            print(f"\n  ✓ PRIMERA FILA VACÍA: {row_idx}")
            print(f"  → Escribir en fila {row_idx}\n")
            break

    wb.close()

except Exception as e:
    print(f"[ERROR] {e}")
