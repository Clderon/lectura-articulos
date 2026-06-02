#!/usr/bin/env python3
"""
Procesador automático de artículos académicos PDF → Excel
Usa DeepSeek como motor de análisis (endpoint compatible OpenAI).
"""

import sys
import os
from pathlib import Path
from dotenv import load_dotenv

# ╔══════════════════════════════════════════════════════════════╗
# ║              CARGAR CONFIGURACIÓN DESDE .env                 ║
# ╠══════════════════════════════════════════════════════════════╣
# ║  Copia .env.example a .env y edita con tus valores          ║
# ║  NUNCA hagas commit del archivo .env (contiene claves)       ║
# ╚══════════════════════════════════════════════════════════════╝

load_dotenv()

# Variables de entorno (obligatorias)
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
PDF_FOLDER = os.getenv("PDF_FOLDER")
EXCEL_PATH = os.getenv("EXCEL_PATH")

# Variables opcionales (con defaults)
MODO_VERIFICACION = int(os.getenv("MODO_VERIFICACION", "0"))
MAX_WORDS_PER_CHUNK = int(os.getenv("MAX_WORDS_PER_CHUNK", "6000"))
SHEET1_NAME = os.getenv("SHEET1_NAME", "Revisión de Artículos")
SHEET2_NAME = os.getenv("SHEET2_NAME", "Lectura Selectiva")

# Validar configuración obligatoria
if not DEEPSEEK_API_KEY:
    print("[ERROR] DEEPSEEK_API_KEY no encontrada en .env")
    print("  Copia .env.example a .env y agrega tu API key")
    sys.exit(1)
if not PDF_FOLDER:
    print("[ERROR] PDF_FOLDER no encontrada en .env")
    sys.exit(1)
if not EXCEL_PATH:
    print("[ERROR] EXCEL_PATH no encontrada en .env")
    sys.exit(1)

# ──────────────────────────────────────────────────────────────

import re
import time
import difflib

# ─────────────────────────── Verificación de dependencias ───────────────────────────

REQUIRED_PACKAGES = {
    "pdfplumber": "pdfplumber",
    "fitz": "pymupdf",
    "openpyxl": "openpyxl",
    "openai": "openai",
    "xlwings": "xlwings",
}

missing_packages = []
for module, package in REQUIRED_PACKAGES.items():
    try:
        __import__(module)
    except ImportError:
        missing_packages.append(package)

if missing_packages:
    print("\n[ERROR] Faltan las siguientes librerías:\n")
    for pkg in missing_packages:
        print(f"  pip install {pkg}")
    print("\nInstálalas y vuelve a ejecutar el script.")
    sys.exit(1)

import pdfplumber
import fitz  # pymupdf
import openpyxl
import xlwings as xw
from openai import OpenAI

# ─────────────────────────── Constantes ───────────────────────────

DEEPSEEK_BASE_URL = "https://api.deepseek.com"
DEEPSEEK_MODEL = "deepseek-chat"
TEMPERATURE = 0.1

# Patrones para cortar antes de la sección de referencias
REFERENCE_STOP_PATTERNS = [
    r"^references$",
    r"^referencias$",
    r"^bibliography$",
    r"^bibliografía$",
    r"^works cited$",
    r"^literatura citada$",
    r"^fuentes$",
    r"^fuentes consultadas$",
]

# Encabezados de sección típicos en artículos académicos
SECTION_HEADER_PATTERNS = [
    r"abstract",
    r"resumen",
    r"keywords?|palabras\s+clave",
    r"introduction|introducción",
    r"background|antecedentes",
    r"related\s+work|trabajos\s+relacionados|estado\s+del\s+arte",
    r"literature\s+review|revisión\s+de\s+literatura",
    r"theoretical\s+framework|marco\s+teórico",
    r"methodology|metodología|method|método",
    r"research\s+design|diseño\s+de\s+investigación",
    r"results?|resultados?",
    r"findings?|hallazgos?",
    r"discussion|discusión",
    r"conclusion|conclusión|conclusions|conclusiones",
    r"future\s+work|trabajo\s+futuro",
    r"limitations?|limitaciones?",
    r"acknowledgments?|agradecimientos?",
]

# ─────────────────────────── Columnas Excel (índices 1-based) ───────────────────────────

# Hoja 1 — "Revisión de Articulos"
# N° | Artículo (Título) | Autor(es) | Base de Datos | Fuente | Nivel Q | SJR | H-Index | N° Citas |
# Pertinencia | Actualidad | Rigor | Fuente (Calidad) | Impacto | Ética | Puntaje | Decisión | Motivo Rechazo F2
COLS_H1 = {
    "N°": 1,
    "Artículo (Título)": 2,
    "Autor(es)": 3,
    # 4: Base de Datos — manual
    # 5: Fuente — manual
    # 6: Nivel Q — manual
    # 7: SJR — manual
    # 8: H-Index — manual
    # 9: N° Citas — manual
    "Pertinencia": 10,
    "Actualidad": 11,
    "Rigor": 12,
    "Fuente (Calidad)": 13,
    "Impacto": 14,
    "Ética": 15,
    "Puntaje": 16,
    "Decisión": 17,
    "Motivo Rechazo F2": 18,
}

# Hoja 2 — "Lectura Selectiva"
# N° | Autor / Año | Objetivo | Método | Hallazgo | Limitación | Uso en tu tesis
COLS_H2 = {
    "N°": 1,
    "Autor / Año": 2,
    "Objetivo": 3,
    "Método": 4,
    "Hallazgo": 5,
    "Limitación": 6,
    "Uso en tu tesis": 7,
}

# ─────────────────────────── Prompt Maestro ───────────────────────────

def build_master_prompt(num1: int, num2: int) -> str:
    return f"""Eres un asistente de investigación especializado en revisiones sistemáticas. Estás apoyando un Estudio de Mapeo Sistemático (SMS) sobre Infrastructure as Code (IaC).

CONTEXTO DEL ESTUDIO:
Título: "Estudio de mapeo sistemático de los enfoques declarativos e imperativos en infraestructura como código: tendencias de investigación, paradigmas y soporte de herramientas"

En este chat analizarás 2 artículos completos. Se te enviarán sección por sección (introducción, metodología, resultados, conclusión, etc.).

INSTRUCCIONES DE RECEPCIÓN:
- Cuando recibas una sección responde SOLO: "Sección [NOMBRE] del Artículo [N°] recibida."
- NO analices ni generes tablas hasta que recibas el mensaje: "CERRAR ARTÍCULO N°[X]"
- Cuando recibas ese mensaje genera las DOS tablas siguientes.

TABLA 1 — RÚBRICA DE VALIDACIÓN
| Campo | Valor |
|---|---|
| N° | |
| Artículo (Título) | |
| Autor(es) | |
| Pertinencia (0-2) | |
| Actualidad (0-2) | |
| Rigor (0-2) | |
| Fuente Calidad (0-2) | |
| Impacto (0-2) | |
| Ética (0-2) | |
| Puntaje Total (/12) | |
| Decisión | |
| Motivo Rechazo F2 | (solo si RECHAZADO F2, sino dejar vacío) |

CRITERIOS DE PUNTUACIÓN:
- Pertinencia (0-2): ¿IaC es el tema principal?
  0 = mención secundaria o contexto
  1 = relevante pero comparte protagonismo
  2 = tema central del artículo
- Actualidad (0-2): año de publicación
  0 = desactualizado
  1 = aceptable pero no reciente
  2 = reciente y vigente
- Rigor (0-2): método, datos, análisis y limitaciones
  0 = ausente o poco claro
  1 = parcial, faltan detalles
  2 = bien descritos
- Fuente Calidad (0-2): revista o conferencia
  0 = no indexada
  1 = indexada pero bajo perfil
  2 = indexada y verificable
- Impacto (0-2): citaciones o relevancia técnica
  0 = sin citaciones ni relevancia
  1 = moderadas o media
  2 = alto número o alta relevancia
- Ética (0-2): transparencia
  0 = sin DOI ni afiliaciones, editorial dudosa
  1 = parciales, editorial aceptable
  2 = DOI, afiliaciones completas, editorial confiable

CRITERIO DE DECISIÓN:
- INCLUIDO: IaC es el OBJETO DE ESTUDIO del artículo. 
  El artículo investiga, analiza, evalúa o compara IaC, 
  sus paradigmas (declarativo, imperativo, híbrido), 
  sus herramientas, prácticas, desafíos o adopción.
  
- RECHAZADO F2: IaC aparece como herramienta de soporte 
  o medio para desplegar otra cosa (una plataforma, 
  un sistema ML, una app, etc.) pero NO es lo que 
  el artículo investiga. El tema principal es otro 
  (ciberseguridad, MLOps, contenedores, etc.) y 
  IaC solo aparece en la implementación.

EXCLUIR SIEMPRE aunque mencionen IaC o enfoque imperativo:
    - Dockerfiles y estudios centrados en Docker como 
    herramienta principal
    - El SMS no considera Dockerfiles como IaC
    - Herramientas IaC válidas: Terraform, Ansible, Puppet, 
    Chef, Pulumi, CloudFormation, OpenTofu, AWS CDK y similares

PREGUNTA CLAVE: ¿El artículo estudia IaC 
o simplemente lo usa para construir otra cosa?
Si lo estudia → INCLUIDO
Si solo lo usa → RECHAZADO F2
  
TABLA 2 — FICHA DEL ARTÍCULO
| Campo | Contenido |
|---|---|
| N° | |
| Autor / Año | |
| Objetivo | |
| Método | |
| Hallazgo | |
| Limitación | |
| Uso en tu tesis | |

CRITERIOS TABLA 2:
- Objetivo: qué problema o pregunta aborda
- Método: cómo lo investiga
- Hallazgo: resultado o conclusión principal
- Limitación: restricciones que el artículo reconoce
- Uso en tu tesis: cómo aporta al SMS, qué enfoque cubre, qué pregunta de investigación responde

IMPORTANTE:
- Genera AMBAS tablas juntas al recibir "CERRAR ARTÍCULO N°[X]"
- No inventes datos que no estén en el texto
- Si un dato no está disponible escribe: No especificado

Responde ahora: "Listo, esperando secciones del Artículo N°{num1} y N°{num2}" """


# ─────────────────────────── Extracción de PDF ───────────────────────────

def extract_with_pdfplumber(pdf_path: str) -> str:
    parts = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text(layout=True)
                if page_text:
                    parts.append(page_text)
                for table in page.extract_tables() or []:
                    for row in table:
                        if row:
                            parts.append(" | ".join(str(c) if c else "" for c in row))
    except Exception as e:
        print(f"    [pdfplumber] Error: {e}")
        return ""
    return "\n".join(parts)


def _find_column_split(text_blocks: list, page_width: float) -> float | None:
    """
    Busca la frontera real entre 2 columnas analizando los espacios vacíos
    en el eje X de todos los bloques de texto.

    Idea: en un PDF de 2 columnas existe una franja vertical vacía (el gutter)
    cerca del centro. La detectamos buscando el gap más ancho en X que esté
    entre el 25% y el 75% del ancho de la página.

    Retorna el X de la mitad del gap, o None si no se detecta layout de 2 cols.
    """
    if not text_blocks:
        return None

    pw = page_width
    # Construir un conjunto de intervalos [x0, x1] ocupados por texto
    # Discretizar en 200 segmentos y marcar cuáles tienen texto
    resolution = 200
    occupied = [False] * resolution

    for b in text_blocks:
        x0 = max(0, int(b[0] / pw * resolution))
        x1 = min(resolution - 1, int(b[2] / pw * resolution))
        for k in range(x0, x1 + 1):
            occupied[k] = True

    # Buscar el gap más largo dentro del rango central (25%–75%)
    lo = int(resolution * 0.25)
    hi = int(resolution * 0.75)
    best_gap_start = -1
    best_gap_len   = 0
    cur_start = -1

    for k in range(lo, hi + 1):
        if not occupied[k]:
            if cur_start < 0:
                cur_start = k
        else:
            if cur_start >= 0:
                gap_len = k - cur_start
                if gap_len > best_gap_len:
                    best_gap_len = gap_len
                    best_gap_start = cur_start
                cur_start = -1

    # Solo considerar 2 columnas si el gap representa al menos el 3% del ancho
    MIN_GAP_FRAC = 0.03
    if best_gap_len < resolution * MIN_GAP_FRAC:
        return None

    gap_mid_frac = (best_gap_start + best_gap_len / 2) / resolution
    return pw * gap_mid_frac


def _extract_page_text(page) -> str:
    """
    Extrae el texto de una página manejando layouts de 1 y 2 columnas.
    Para 2 columnas: lee la columna izquierda completa, luego la derecha.
    Para 1 columna: lee en orden natural de arriba a abajo.
    Bloques que abarcan todo el ancho (títulos, abstracts) se colocan primero.
    """
    pw = page.rect.width
    blocks = page.get_text("blocks")
    text_blocks = [b for b in blocks if b[6] == 0 and b[4].strip()]

    if not text_blocks:
        return ""

    split_x = _find_column_split(text_blocks, pw)

    if split_x is None:
        # Una sola columna — orden por Y
        return "\n".join(b[4] for b in sorted(text_blocks, key=lambda b: b[1]))

    # Dos columnas detectadas
    # Bloques que abarcan el punto de separación = encabezados / pie de página de ancho completo
    full_width = sorted(
        [b for b in text_blocks if b[0] < split_x * 0.5 and b[2] > split_x * 1.5],
        key=lambda b: b[1]
    )
    full_set = set(id(b) for b in full_width)
    # Asignar por el CENTRO del bloque — evita que bloques cercanos al gutter
    # queden en la columna equivocada
    col_left = sorted(
        [b for b in text_blocks if id(b) not in full_set and (b[0] + b[2]) / 2 < split_x],
        key=lambda b: b[1]
    )
    col_right = sorted(
        [b for b in text_blocks if id(b) not in full_set and (b[0] + b[2]) / 2 >= split_x],
        key=lambda b: b[1]
    )

    parts = (
        [b[4] for b in full_width if b[1] < (page.rect.height * 0.15)]  # cabecera de página
        + [b[4] for b in col_left]
        + [b[4] for b in col_right]
        + [b[4] for b in full_width if b[1] >= (page.rect.height * 0.15)]  # pie de página
    )
    return "\n".join(parts)


def extract_with_pymupdf(pdf_path: str) -> str:
    parts = []
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            text = _extract_page_text(page)
            if text:
                parts.append(text)
        doc.close()
    except Exception as e:
        print(f"    [pymupdf] Error: {e}")
        return ""
    return "\n".join(parts)


def extract_pdf_text(pdf_path: str) -> str:
    """
    Extrae el texto del PDF. Usa pymupdf como principal porque maneja mejor
    los PDFs de 2 columnas (separa columnas por posición). pdfplumber como fallback.
    """
    print("    Extrayendo con pymupdf (manejo de columnas)...")
    text = extract_with_pymupdf(pdf_path)
    if len(text.strip()) < 200:
        print("    Texto insuficiente, usando pdfplumber como fallback...")
        text = extract_with_pdfplumber(pdf_path)
    if len(text.strip()) < 200:
        raise ValueError("No se pudo extraer texto legible del PDF.")
    return text


# Secciones que vale la pena rescatar aunque aparezcan después de REFERENCES
_RESCUE_AFTER_REFS = re.compile(
    r"^(?:[IVXLCDM]{1,6}[.\s]+|[A-Z]\.\s+)?"
    r"(?:conclusion|conclusi[oó]n|conclusiones|summary|resumen|"
    r"acknowledgment|acknowledgement|agradecimiento|future\s+work|trabajo\s+futuro)",
    re.IGNORECASE
)


def cut_at_references(text: str) -> str:
    """
    Corta antes de la sección de referencias buscando solo en el último 30%.
    Después del corte, rescata cualquier sección real (conclusión, etc.) que
    haya quedado enterrada en el texto de referencias — esto ocurre en PDFs de
    2 columnas donde la columna de referencias y la de contenido se intercalan.
    """
    lines = text.split("\n")
    total = len(lines)
    search_from = int(total * 0.70)

    cut_line = None
    for i, line in enumerate(lines):
        if i < search_from:
            continue
        stripped = line.strip().lower()
        for pattern in REFERENCE_STOP_PATTERNS:
            if re.fullmatch(pattern, stripped, re.IGNORECASE):
                cut_line = i
                break
        if cut_line is not None:
            break

    if cut_line is None:
        return text

    kept = lines[:cut_line]
    discarded = lines[cut_line + 1:]

    # Rescatar secciones reales que quedaron mezcladas con las referencias
    rescued = []
    in_rescue = False
    for line in discarded:
        s = line.strip()
        if s and _RESCUE_AFTER_REFS.match(s):
            in_rescue = True
        if in_rescue:
            rescued.append(line)

    if rescued:
        return "\n".join(kept + [""] + rescued)
    return "\n".join(kept)


# Patrones que indican que una línea NO es encabezado de sección
_NOT_HEADER_RE = re.compile(
    r"[•◦·]{2,}"
    r"|[–—]{2,}"
    r"|\w+\s*et\s+al\b"
    r"|https?://"
    r"|www\."
    r"|\w{3,}\.\w+/\w+"        # Autor./Revista — mínimo 3 chars antes del punto (evita atrapar "C.")
    r"|\(\s*[•\d]{4,}\s*\)"
    r"|^\d+\s+\w+\.\w+\."
    r"|@|\bDOI\b|doi:"
    , re.IGNORECASE
)


def _title_portion(s: str) -> str:
    """
    En PDFs de 2 columnas el encabezado y el texto de la columna opuesta
    quedan en la misma línea separados por 3+ espacios.
    Devuelve solo la parte izquierda (el título real).
    Ej: "I. INTRODUCTION           deployment actions." → "I. INTRODUCTION"
    """
    m = re.match(r"^(.{3,70}?)\s{3,}\S", s)
    return m.group(1).strip() if m else s.strip()


def _is_section_header(line: str) -> bool:
    """
    Detecta encabezados de sección. Soporta:
    - Arábigos:  1.  2.1  4.2.1  Título
    - Romanos:   I.  II.  IV. CONCLUSION
    - Letras:    A.  B.   C. Background  (subsecciones IEEE/ACM)
    - CAPS:      ABSTRACT  CONCLUSION  (≥1 palabra conocida o ≥2 palabras)
    - Nombres conocidos con prefijo opcional
    - PDFs 2 columnas: evalúa solo la porción izquierda de la línea
    """
    raw = line.strip()
    if len(raw) < 3:
        return False

    # Trabajar sobre la porción de título (descarta texto de columna opuesta)
    s = _title_portion(raw)
    if len(s) > 150:
        return False

    # Excluir leyendas de figuras/tablas
    if re.match(
        r"^(fig\.?|figure|tabla?|table|listing|algorithm|ecuaci[oó]n|equation|appendix)\s*[\d:]",
        s, re.IGNORECASE
    ):
        return False

    # Excluir encabezados de página / notas al pie / URLs
    if _NOT_HEADER_RE.search(s):
        return False

    known = "|".join(SECTION_HEADER_PATTERNS)

    # Regla 1 — arábigos: "1.", "2.1", "4.2.1  Título"
    if re.match(r"^\d[\d.]*[\s.]\s*\S", s):
        rest = re.sub(r"^\d[\d.\s]*", "", s).strip()
        rest = _title_portion(rest)

        # Rechazar si parece mitad de una oración partida por salto de línea:
        # - Termina con preposición, conjunción o artículo (frase incompleta)
        # - Contiene verbo conjugado típico de prosa ("is", "are", "was", "can", etc.)
        _incomplete_end = re.compile(
            r"\b(to|a|an|the|of|in|at|by|for|with|is|are|was|were|as|"
            r"and|or|but|if|when|that|this|these|those|it|its|be|been)\s*$",
            re.IGNORECASE
        )
        _prose_verb = re.compile(
            r"\b(is|are|was|were|has|have|had|can|will|should|does|do|"
            r"configured|defined|applied|used|created|called|named|shown)\b",
            re.IGNORECASE
        )
        if _incomplete_end.search(rest) or _prose_verb.search(rest):
            # Solo permitir si coincide exactamente con nombre de sección conocido
            if not re.search(rf"\b({known})\b", s, re.IGNORECASE):
                return False

        if rest and not rest[0].islower() and len(rest.split()) <= 12:
            return True
        if re.search(rf"\b({known})\b", s, re.IGNORECASE):
            return True
        return False

    # Regla 2 — romanos: "I.", "II.", "IV. CONCLUSION"
    if re.match(r"^[IVX]{1,6}[\s.]+\S", s):
        rest = re.sub(r"^[IVX]+[\s.]+", "", s).strip()
        rest = _title_portion(rest)
        if rest and not rest[0].islower() and len(rest.split()) <= 12:
            return True
        if re.search(rf"\b({known})\b", s, re.IGNORECASE):
            return True
        return False

    # Regla 3 — letra sola: "A. Background", "B. Results"
    if re.match(r"^[A-Z]\.\s+[A-Za-z]", s):
        rest = re.sub(r"^[A-Z]\.\s+", "", s).strip()
        if len(rest.split()) <= 10 and not rest[0].islower():
            return True

    # Regla 4 — TODO MAYÚSCULAS
    words_s = s.split()
    if s == s.upper() and re.search(r"[A-Z]{2,}", s):
        # Una palabra conocida es suficiente (ABSTRACT, CONCLUSION, etc.)
        if re.search(rf"\b({known})\b", s, re.IGNORECASE):
            return True
        # O mínimo 2 palabras caps cortas
        if 2 <= len(words_s) <= 12:
            return True

    # Regla 5 — nombres conocidos: el título limpio debe ser SOLO el nombre de sección
    # (sin texto de prosa después). Usa _title_portion para manejar PDFs 2 columnas.
    title_clean = _title_portion(s)
    if re.fullmatch(rf"(?:[\d\s.ivxIVX]{{0,10}})?({known})[\s.]*", title_clean, re.IGNORECASE):
        return True

    return False


# Líneas que son ruido de página y deben eliminarse del contenido
_PAGE_NOISE_RE = re.compile(
    r"[•◦·]{3,}"                     # puntos decorativos •••
    r"|\w+et\s+al\b.*\d{4}"          # cita estilo "Rahman et al. 2021"
    r"|\w+\.\w+\./[A-Za-z]"          # patrón Autor./Revista
    r"|^\s*\d{1,3}\s*$"              # página sola (solo un número)
    , re.IGNORECASE | re.MULTILINE
)


def _clean_content(text: str) -> str:
    """Elimina líneas que son encabezados/pies de página del contenido de la sección."""
    lines = []
    for line in text.split("\n"):
        s = line.strip()
        # Saltar líneas que son claramente ruido de encabezado de página
        if _NOT_HEADER_RE.search(s) and len(s) < 120:
            continue
        lines.append(line)
    return "\n".join(lines)


# ── Utilidades para detección de secuencias numéricas ──────────────────────

_ROMAN_MAP = [
    (10,'X'),(9,'IX'),(5,'V'),(4,'IV'),
    (1,'I')
]
_ROMAN_TO_INT = {
    'I':1,'II':2,'III':3,'IV':4,'V':5,
    'VI':6,'VII':7,'VIII':8,'IX':9,'X':10,
    'XI':11,'XII':12,'XIII':13,'XIV':14,'XV':15,
}

def _roman_to_int(s: str) -> int:
    return _ROMAN_TO_INT.get(s.upper().strip(), 0)

def _int_to_roman(n: int) -> str:
    result = ""
    for val, sym in _ROMAN_MAP:
        while n >= val:
            result += sym
            n -= val
    return result

def _get_prefix(name: str):
    """
    Extrae el tipo y valor del prefijo de numeración de un encabezado.
    Retorna ('roman', int) | ('letter', str) | ('arabic', int) | None
    """
    s = name.strip()
    # Romano: I., II., IV. TITLE
    m = re.match(r'^([IVXLCDM]{1,6})[.\s]', s)
    if m:
        val = _roman_to_int(m.group(1))
        if 1 <= val <= 20:
            return ('roman', val)
    # Letra: A. Title
    m = re.match(r'^([A-Z])\.\s', s)
    if m:
        return ('letter', m.group(1))
    # Arábigo simple: 1. Title  (no subsecciones 1.1)
    m = re.match(r'^(\d+)\.\s+[A-Za-z]', s)
    if m and not re.match(r'^\d+\.\d', s):
        return ('arabic', int(m.group(1)))
    return None


def _next_in_sequence(prefix: tuple):
    """Devuelve el siguiente valor esperado en la secuencia."""
    seq_type, val = prefix
    if seq_type == 'letter':
        return (seq_type, chr(ord(val) + 1))
    return (seq_type, val + 1)


def _prefix_to_str(prefix: tuple) -> str:
    seq_type, val = prefix
    if seq_type == 'roman':
        return _int_to_roman(val)
    return str(val)


def _search_header_in_content(content: str, search_str: str):
    """
    Busca un encabezado 'search_str.' al inicio de línea dentro del contenido.
    Retorna el match o None.
    """
    pat = re.compile(
        rf"(?m)^({re.escape(search_str)}[.\s]{{1,3}}\S.{{0,100}})$"
    )
    return pat.search(content)


def _recover_sequence_gaps(sections_list: list) -> list:
    """
    Segunda pasada: recupera secciones cuyos encabezados no fueron detectados.

    Para cada sección con prefijo de secuencia (A., I., 1.) busca el SIGUIENTE
    elemento esperado en DOS lugares:
      1. Dentro del PROPIO contenido (el header faltante quedó ahí)
      2. Dentro del contenido de la SIGUIENTE sección detectada (si hubo merge)

    Itera hasta 5 veces para resolver gaps encadenados (A→D con B y C faltantes).
    """
    parsed = [{'name': n, 'content': c, 'prefix': _get_prefix(n)}
              for n, c in sections_list]

    for _ in range(5):
        new_parsed = []
        added = False
        i = 0

        while i < len(parsed):
            item = parsed[i]
            p = item['prefix']

            if p:
                nxt_prefix = _next_in_sequence(p)
                nxt_str = _prefix_to_str(nxt_prefix)

                # ¿El siguiente ya fue detectado?
                already_exists = any(
                    x['prefix'] == nxt_prefix for x in parsed
                )

                if not already_exists:
                    # Búsqueda 1: en el contenido PROPIO (el header quedó ahí sin detectar)
                    m = _search_header_in_content(item['content'], nxt_str)
                    if m:
                        sp = m.start()
                        rec_name = _title_portion(m.group(1).strip())
                        new_parsed.append({**item, 'content': item['content'][:sp].strip()})
                        # El resto se convierte en la sección recuperada
                        recovered = {
                            'name': rec_name,
                            'content': item['content'][m.end():].strip(),
                            'prefix': nxt_prefix,
                        }
                        # Insertar antes de procesar la siguiente sección real
                        # (necesitamos seguir iterando sobre la recuperada)
                        parsed.insert(i + 1, recovered)
                        added = True
                        i += 1
                        continue

                    # Búsqueda 2: en el contenido de la SIGUIENTE sección detectada
                    # (ocurre cuando el header faltante fue absorbido por merge en la siguiente)
                    if i + 1 < len(parsed):
                        nxt_item = parsed[i + 1]
                        m2 = _search_header_in_content(nxt_item['content'], nxt_str)
                        if m2:
                            sp2 = m2.start()
                            rec_name2 = _title_portion(m2.group(1).strip())
                            # El contenido antes del header recuperado pertenece a la sección faltante
                            # (estaba mezclado al inicio del siguiente bloque)
                            before = nxt_item['content'][:sp2].strip()
                            after  = nxt_item['content'][m2.end():].strip()
                            # Actualizar siguiente sección para que solo tenga su contenido real
                            parsed[i + 1] = {**nxt_item, 'content': after}
                            new_parsed.append(item)
                            new_parsed.append({
                                'name': rec_name2,
                                'content': before,
                                'prefix': nxt_prefix,
                            })
                            added = True
                            i += 1
                            continue

            new_parsed.append(item)
            i += 1

        parsed = new_parsed
        if not added:
            break

    return [(item['name'], item['content']) for item in parsed]


def split_into_sections(text: str) -> dict:
    """
    Divide el texto en secciones detectando encabezados por:
    numeración, todo mayúsculas, y nombres conocidos.
    Las secciones muy cortas (< 40 palabras) se fusionan con la siguiente
    para evitar fragmentos de encabezados multi-línea.
    """
    lines = text.split("\n")
    sections_raw = []        # lista de (nombre, contenido)
    current_name = "Inicio"
    current_lines = []

    for line in lines:
        stripped = line.strip()
        if stripped and _is_section_header(stripped):
            if current_lines:
                sections_raw.append((current_name, "\n".join(current_lines).strip()))
            # Usar solo la porción de título (sin texto de columna opuesta)
            header_name = _title_portion(stripped)

            # Caso "Abstract—contenido" o "Abstract: contenido" — separar en el separador
            dash_match = re.match(r"^([A-Za-z ]{3,30})[—–:]\s*(.+)", header_name)
            if dash_match and len(dash_match.group(1).split()) <= 3:
                header_name = dash_match.group(1).strip()
                extra = dash_match.group(2).strip()
            else:
                extra = stripped[len(header_name):].strip() if len(stripped) > len(header_name) else ""

            current_name = header_name
            current_lines = [extra] if extra else []
        else:
            current_lines.append(line)

    if current_lines:
        sections_raw.append((current_name, "\n".join(current_lines).strip()))

    # Limpiar ruido de encabezados/pies de página del contenido
    sections_raw = [(n, _clean_content(c)) for n, c in sections_raw]

    # Fusionar secciones con contenido muy corto (encabezados multi-línea o falsas detecciones)
    # Regla: si la sección tiene un prefijo de secuencia válido (A., I., 1.) solo se fusiona
    # si tiene < 8 palabras (casi vacía = header partido). El resto usa el umbral normal.
    MIN_WORDS_GENERIC  = 30   # secciones sin prefijo de secuencia
    MIN_WORDS_SEQUENCE = 0    # secciones con prefijo (A., I., 1.) — NUNCA fusionar por contenido vacío
    merged = {}
    i = 0
    while i < len(sections_raw):
        name, content = sections_raw[i]
        word_count = len(content.split())
        threshold = MIN_WORDS_SEQUENCE if _get_prefix(name) else MIN_WORDS_GENERIC
        if word_count < threshold and i + 1 < len(sections_raw):
            next_name, next_content = sections_raw[i + 1]
            combined_name = f"{name} {next_name}" if word_count < 5 else next_name
            sections_raw[i + 1] = (combined_name, (content + "\n" + next_content).strip())
            i += 1
            continue
        # Guardar siempre si tiene contenido, O si es sección con prefijo de secuencia
        # (II. OVERVIEW puede tener 0 palabras — es válida como encabezado estructural)
        if content or _get_prefix(name) is not None:
            merged[name] = content
        i += 1

    if not merged:
        return {"Contenido Completo": text}

    # Segunda pasada: recuperar secciones cuyos headers quedaron ocultos en el contenido
    recovered_list = _recover_sequence_gaps(list(merged.items()))

    # Reconstruir como dict preservando orden
    final = {}
    for name, content in recovered_list:
        if content or _get_prefix(name) is not None:
            final[name] = content

    return final if final else {"Contenido Completo": text}


def chunk_text(text: str, max_words: int) -> list:
    words = text.split()
    return [
        " ".join(words[i:i + max_words])
        for i in range(0, len(words), max_words)
    ]


# ─────────────────────────── Parser de tablas markdown ───────────────────────────

def parse_markdown_table(block: str) -> dict:
    result = {}
    for line in block.strip().split("\n"):
        line = line.strip()
        if not line.startswith("|"):
            continue
        parts = [p.strip() for p in line.split("|") if p.strip()]
        if len(parts) < 2:
            continue
        key, value = parts[0], parts[1]
        if re.fullmatch(r"[-: ]+", value):
            continue
        if key.lower() in ("campo", "---"):
            continue
        result[key] = value
    return result


def parse_tables_from_response(response: str, article_number: int) -> tuple:
    """Extrae tabla1 (rúbrica) y tabla2 (ficha) del texto de respuesta."""
    table_blocks = re.findall(r"(\|[^\n]+(?:\n\|[^\n]+)+)", response)

    tabla1 = {}
    tabla2 = {}

    for block in table_blocks:
        parsed = parse_markdown_table(block)
        keys_lower = {k.lower() for k in parsed}
        if any(k in keys_lower for k in ("artículo (título)", "artículo", "pertinencia (0-2)", "pertinencia")):
            tabla1 = parsed
        elif any(k in keys_lower for k in ("autor / año", "autor/año", "objetivo", "hallazgo")):
            tabla2 = parsed

    # Fallback por posición si no se identificaron por contenido
    if not tabla1 and len(table_blocks) >= 1:
        tabla1 = parse_markdown_table(table_blocks[0])
    if not tabla2 and len(table_blocks) >= 2:
        tabla2 = parse_markdown_table(table_blocks[1])

    return tabla1, tabla2


# ─────────────────────────── Normalización de claves ───────────────────────────

TABLE1_KEY_MAP = {
    "n°": "N°",
    "artículo (título)": "Artículo (Título)",
    "artículo": "Artículo (Título)",
    "titulo": "Artículo (Título)",
    "título": "Artículo (Título)",
    "autor(es)": "Autor(es)",
    "autores": "Autor(es)",
    "autor": "Autor(es)",
    "pertinencia (0-2)": "Pertinencia",
    "pertinencia": "Pertinencia",
    "actualidad (0-2)": "Actualidad",
    "actualidad": "Actualidad",
    "rigor (0-2)": "Rigor",
    "rigor": "Rigor",
    "fuente calidad (0-2)": "Fuente (Calidad)",
    "fuente (calidad)": "Fuente (Calidad)",
    "fuente calidad": "Fuente (Calidad)",
    "impacto (0-2)": "Impacto",
    "impacto": "Impacto",
    "ética (0-2)": "Ética",
    "ética": "Ética",
    "etica (0-2)": "Ética",
    "etica": "Ética",
    "puntaje total (/12)": "Puntaje",
    "puntaje total": "Puntaje",
    "puntaje": "Puntaje",
    "decisión": "Decisión",
    "decision": "Decisión",
    "motivo rechazo f2": "Motivo Rechazo F2",
    "motivo de rechazo f2": "Motivo Rechazo F2",
}

TABLE2_KEY_MAP = {
    "n°": "N°",
    "autor / año": "Autor / Año",
    "autor/año": "Autor / Año",
    "autor": "Autor / Año",
    "objetivo": "Objetivo",
    "método": "Método",
    "metodo": "Método",
    "hallazgo": "Hallazgo",
    "hallazgos": "Hallazgo",
    "limitación": "Limitación",
    "limitacion": "Limitación",
    "limitaciones": "Limitación",
    "uso en tu tesis": "Uso en tu tesis",
    "uso en la tesis": "Uso en tu tesis",
}


def normalize_table(raw: dict, key_map: dict, article_number: int) -> dict:
    normalized = {}
    for k, v in raw.items():
        norm_key = key_map.get(k.lower(), k)
        normalized[norm_key] = v
    if "N°" not in normalized:
        normalized["N°"] = str(article_number)
    return normalized


# ─────────────────────────── DeepSeek API ───────────────────────────

def send_to_deepseek(client: OpenAI, messages: list, content: str, retries: int = 3) -> str:
    messages.append({"role": "user", "content": content})
    last_error = None

    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=DEEPSEEK_MODEL,
                messages=messages,
                temperature=TEMPERATURE,
            )
            reply = response.choices[0].message.content
            messages.append({"role": "assistant", "content": reply})
            return reply
        except Exception as e:
            last_error = e
            print(f"    [DeepSeek] Intento {attempt + 1}/{retries} falló: {e}")
            if attempt < retries - 1:
                wait = 5 * (attempt + 1)
                print(f"    Esperando {wait}s antes de reintentar...")
                time.sleep(wait)

    raise RuntimeError(f"DeepSeek no respondió tras {retries} intentos. Último error: {last_error}")


# ─────────────────────────── Preview de secciones ───────────────────────────

EXPORT_DIR = Path(__file__).parent / "secciones_exportadas"


def seccion_preview_e_interaccion(pdf_name: str, sections: dict):
    """
    Lista las secciones detectadas con su número de palabras.
    Si MODO_VERIFICACION está activo, permite al usuario elegir
    una sección por número y la exporta a un archivo .txt para
    compararla con el PDF original.
    Puede repetirse: el usuario escribe números hasta que presione Enter.
    """
    sec_list = list(sections.items())
    total_words = sum(len(c.split()) for c in sections.values())

    print(f"\n  {'─'*60}")
    print(f"  SECCIONES DETECTADAS — {pdf_name}")
    print(f"  {len(sec_list)} secciones  |  {total_words} palabras en total")
    print(f"  {'─'*60}")
    for i, (name, content) in enumerate(sec_list, 1):
        words = len(content.split())
        print(f"    {i:>2}.  [{words:>5} pal]  {name}")
    print(f"  {'─'*60}")

    if not MODO_VERIFICACION:
        return

    EXPORT_DIR.mkdir(exist_ok=True)
    print("  [MODO VERIFICACIÓN]")
    print("  Escribe el número de una sección para exportarla a .txt y compararla")
    print("  con el documento. Puedes exportar varias. Enter para continuar.\n")

    while True:
        raw = input("  Número de sección (o Enter para continuar): ").strip()
        if not raw:
            break
        if not raw.isdigit():
            print("  Ingresa un número válido.")
            continue
        idx = int(raw) - 1
        if idx < 0 or idx >= len(sec_list):
            print(f"  Número fuera de rango (1-{len(sec_list)}).")
            continue

        sec_name, sec_content = sec_list[idx]
        safe_name = re.sub(r"[^\w\s-]", "", sec_name)[:40].strip().replace(" ", "_")
        pdf_stem = Path(pdf_name).stem[:25]
        out_file = EXPORT_DIR / f"{pdf_stem}__{safe_name}.txt"

        with open(out_file, "w", encoding="utf-8") as f:
            f.write(f"ARTÍCULO : {pdf_name}\n")
            f.write(f"SECCIÓN  : {sec_name}\n")
            f.write(f"PALABRAS : {len(sec_content.split())}\n")
            f.write("=" * 60 + "\n\n")
            f.write(sec_content)

        print(f"  → Exportado: {out_file}\n")


def process_article(client: OpenAI, messages: list, art_num: int, pdf_path: Path, total: int, current: int) -> tuple:
    """
    Envía las secciones de un artículo a DeepSeek y retorna (tabla1, tabla2).
    """
    print(f"\n  ── Artículo N°{art_num}: {pdf_path.name}")
    print(f"  Procesando artículo {current} de {total}")

    # Extracción de texto
    raw_text = extract_pdf_text(str(pdf_path))
    raw_text = cut_at_references(raw_text)
    sections = split_into_sections(raw_text)

    # Mostrar secciones e interacción de verificación
    seccion_preview_e_interaccion(pdf_path.name, sections)

    # Agrupar secciones muy cortas con la siguiente para no enviar mensajes vacíos.
    # Una sección con prefijo de secuencia (II., A.) y < 30 palabras se combina
    # con la que le sigue — se preserva el nombre para que DeepSeek vea la estructura.
    sections_to_send = []
    sec_items = list(sections.items())
    i = 0
    while i < len(sec_items):
        name, content = sec_items[i]
        word_count = len(content.split())
        if word_count < 30 and _get_prefix(name) is not None and i + 1 < len(sec_items):
            next_name, next_content = sec_items[i + 1]
            combined_label = f"{name} / {next_name}"
            combined_content = (content + "\n" + next_content).strip()
            sections_to_send.append((combined_label, combined_content))
            i += 2
        else:
            sections_to_send.append((name, content))
            i += 1

    # Enviar secciones
    total_sections = len(sections_to_send)
    for sec_idx, (sec_name, sec_content) in enumerate(sections_to_send, 1):
        chunks = chunk_text(sec_content, MAX_WORDS_PER_CHUNK)
        for chunk_idx, chunk in enumerate(chunks, 1):
            part_tag = f" (parte {chunk_idx}/{len(chunks)})" if len(chunks) > 1 else ""
            label = f"  Enviando '{sec_name}'{part_tag} [{sec_idx}/{total_sections}]..."
            print(label)
            msg = f"SECCIÓN: {sec_name}{part_tag}\nARTÍCULO N°{art_num}\n\n{chunk}"
            reply = send_to_deepseek(client, messages, msg)
            if "recibida" not in reply.lower():
                print(f"    [AVISO] Respuesta inesperada: {reply[:120]}")

    # Cerrar artículo y recibir tablas
    print(f"  Enviando CERRAR ARTÍCULO N°{art_num}...")
    final_reply = send_to_deepseek(client, messages, f"CERRAR ARTÍCULO N°{art_num}")

    # Parsear y normalizar tablas
    raw_t1, raw_t2 = parse_tables_from_response(final_reply, art_num)
    tabla1 = normalize_table(raw_t1, TABLE1_KEY_MAP, art_num)
    tabla2 = normalize_table(raw_t2, TABLE2_KEY_MAP, art_num)

    return tabla1, tabla2


# ─────────────────────────── Excel ───────────────────────────

def find_target_row(sheet, article_number: int, num_col: int = 1) -> int:
    """
    Busca dónde escribir el artículo:
      - Si ya existe una fila con ese N° → usa esa fila (actualiza)
      - Si hay una fila con ese N° pero sin contenido → usa esa fila (placeholder vacío)
      - Si no existe → agrega después de la última fila con N° Y contenido real
    Una fila con solo N° y sin título/autor se considera vacía.
    Detecta automáticamente la fila de cabecera para no escribir en ella.
    """
    data_start = find_header_row(sheet) + 1
    last_content_row = data_start - 1
    placeholder_row = None

    for row in sheet.iter_rows(min_row=data_start):
        num_val = row[num_col - 1].value
        if num_val is None:
            continue

        # Verificar si la fila tiene contenido real (columna siguiente al N°)
        next_val = row[num_col].value if len(row) > num_col else None
        has_content = next_val is not None and str(next_val).strip() != ""

        if has_content:
            last_content_row = row[num_col - 1].row

        try:
            if int(str(num_val).strip()) == article_number:
                if has_content:
                    return row[num_col - 1].row   # fila existente con datos
                else:
                    placeholder_row = row[num_col - 1].row  # placeholder vacío
        except (ValueError, TypeError):
            pass

    if placeholder_row is not None:
        return placeholder_row  # reusar el placeholder

    return last_content_row + 1  # agregar al final del contenido real


def parse_score(value) -> object:
    """Extrae el primer número de un string o lo devuelve tal cual."""
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else str(value)


def write_article_to_excel(excel_path: str, art_num: int, tabla1: dict, tabla2: dict):
    """
    Escribe los datos en el Excel usando xlwings (COM de Excel).
    Preserva todo el formato, tablas y funcionalidades del archivo original.
    """
    # Determinar filas destino con openpyxl en modo solo lectura (rápido, sin modificar)
    wb_r = openpyxl.load_workbook(excel_path, read_only=True)
    row1 = find_target_row(wb_r[SHEET1_NAME], art_num)
    row2 = find_target_row(wb_r[SHEET2_NAME], art_num)
    wb_r.close()

    # Escribir a través del COM de Excel (sin corrupción de formato)
    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(excel_path)

        def w(sheet, row, col, value):
            sheet.cells(row, col).value = value

        s1 = wb.sheets[SHEET1_NAME]
        w(s1, row1, COLS_H1["N°"],               art_num)
        w(s1, row1, COLS_H1["Artículo (Título)"], tabla1.get("Artículo (Título)", ""))
        w(s1, row1, COLS_H1["Autor(es)"],          tabla1.get("Autor(es)", ""))
        w(s1, row1, COLS_H1["Pertinencia"],        parse_score(tabla1.get("Pertinencia")))
        w(s1, row1, COLS_H1["Actualidad"],         parse_score(tabla1.get("Actualidad")))
        w(s1, row1, COLS_H1["Rigor"],              parse_score(tabla1.get("Rigor")))
        w(s1, row1, COLS_H1["Fuente (Calidad)"],   parse_score(tabla1.get("Fuente (Calidad)")))
        w(s1, row1, COLS_H1["Impacto"],            parse_score(tabla1.get("Impacto")))
        w(s1, row1, COLS_H1["Ética"],              parse_score(tabla1.get("Ética")))
        w(s1, row1, COLS_H1["Puntaje"],            parse_score(tabla1.get("Puntaje")))
        w(s1, row1, COLS_H1["Decisión"],           tabla1.get("Decisión", ""))
        w(s1, row1, COLS_H1["Motivo Rechazo F2"],  tabla1.get("Motivo Rechazo F2", ""))

        s2 = wb.sheets[SHEET2_NAME]
        w(s2, row2, COLS_H2["N°"],              art_num)
        w(s2, row2, COLS_H2["Autor / Año"],     tabla2.get("Autor / Año", ""))
        w(s2, row2, COLS_H2["Objetivo"],        tabla2.get("Objetivo", ""))
        w(s2, row2, COLS_H2["Método"],          tabla2.get("Método", ""))
        w(s2, row2, COLS_H2["Hallazgo"],        tabla2.get("Hallazgo", ""))
        w(s2, row2, COLS_H2["Limitación"],      tabla2.get("Limitación", ""))
        w(s2, row2, COLS_H2["Uso en tu tesis"], tabla2.get("Uso en tu tesis", ""))

        # Si la decisión es RECHAZADO, marcar el N° con fondo rojo en Hoja 2
        decision = tabla1.get("Decisión", "").strip()
        if decision.upper() == "RECHAZADO" or decision.upper().startswith("RECHAZADO"):
            cell_num = s2.cells(row2, COLS_H2["N°"])
            cell_num.color = (255, 0, 0)  # RGB rojo

        wb.save()
        wb.close()
        print(f"  [Excel guardado] N°{art_num} → Hoja1 fila {row1}, Hoja2 fila {row2}")

    except Exception as e:
        raise RuntimeError(f"Error al escribir en Excel con xlwings: {e}")
    finally:
        app.quit()


# ─────────────────────────── Artículos ya analizados en Excel ───────────────────────────

def find_header_row(ws) -> int:
    """
    Encuentra la fila de cabecera buscando 'N°' en la columna A.
    Los datos reales empiezan en header_row + 1.
    """
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=1):
        val = row[0].value
        if val is not None and str(val).strip() in ("N°", "N", "Nro", "No.", "Número"):
            return row[0].row
    return 1  # fallback: la cabecera es la fila 1


def load_analyzed_articles(excel_path: str) -> tuple:
    """
    Lee el Excel y retorna:
      - analyzed: dict {titulo_normalizado: (N°, titulo_original)}
      - last_num: último N° cuya fila tiene cualquier contenido real (no solo el N°).

    Una fila cuenta como "analizada" si tiene N° y AL MENOS UNO de estos no vacío:
      - Título (col 2)
      - Autor (col 3)
      - Decisión (col 17)
    Esto cubre el caso de filas con título vacío pero con otros datos ya cargados.
    """
    analyzed = {}
    last_num = 0
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[SHEET1_NAME]
        data_start = find_header_row(ws) + 1
        for row in ws.iter_rows(min_row=data_start, max_col=17, values_only=True):
            num_val = row[0]  # col 1: N°
            if num_val is None:
                continue
            try:
                n = int(str(num_val).strip())
            except (ValueError, TypeError):
                continue

            title_val  = row[1]   # col 2: Título
            autor_val  = row[2]   # col 3: Autor
            dec_val    = row[16]  # col 17: Decisión

            title_str = str(title_val).strip() if title_val else ""
            autor_str = str(autor_val).strip() if autor_val else ""
            dec_str   = str(dec_val).strip()   if dec_val   else ""

            # La fila tiene contenido real si cualquiera de los campos clave está lleno
            has_content = bool(title_str or autor_str or dec_str)

            if has_content:
                if n > last_num:
                    last_num = n
                # Para la comparación de duplicados usamos el título si existe,
                # si no, usamos un placeholder con el N° para al menos contar la fila
                key_title = title_str if title_str else f"[Artículo N°{n} sin título]"
                analyzed[_normalize_title(key_title)] = (n, key_title)

        wb.close()
    except Exception as e:
        print(f"  [AVISO] No se pudo leer el Excel: {e}")
    return analyzed, last_num


def _normalize_title(title: str) -> str:
    """Normaliza un título para comparación: minúsculas, sin puntuación extra."""
    t = title.lower().strip()
    t = re.sub(r"[^\w\s]", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t


def extract_quick_title(pdf_path) -> list:
    """
    Retorna una lista de candidatos de título ordenados por confianza:
      1. Metadata del PDF (más confiable si existe)
      2. Primeras líneas sustanciales de la primera página
      3. Nombre del archivo normalizado (siempre disponible como fallback)
    """
    candidates = []

    # 1. Metadata del PDF
    try:
        doc = fitz.open(str(pdf_path))
        meta_title = (doc.metadata.get("title", "") or "").strip()
        doc.close()
        if len(meta_title) > 20:
            candidates.append(meta_title)
    except Exception:
        pass

    # 2. Primera página — líneas sustanciales
    page_text = ""
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            if pdf.pages:
                page_text = pdf.pages[0].extract_text(layout=False) or ""
    except Exception:
        try:
            doc = fitz.open(str(pdf_path))
            page_text = doc[0].get_text("text") if doc.page_count > 0 else ""
            doc.close()
        except Exception:
            pass

    if page_text:
        skip_re = re.compile(
            r"^\s*("
            r"doi\s*:|issn|isbn|©|copyright|vol\.|volume|issue|pages?|pp\.|journal|"
            r"proceedings|conference|abstract|keywords?|received|accepted|published|"
            r"https?://|www\.|email|@|\d{4}[-/]\d{2}[-/]\d{2}|ieee|acm|springer|elsevier"
            r")",
            re.IGNORECASE
        )
        page_candidates = []
        for line in page_text.split("\n"):
            line = line.strip()
            if len(line) < 15:
                continue
            if skip_re.match(line):
                continue
            if re.fullmatch(r"[\d\s\W]+", line):
                continue
            page_candidates.append(line)
            if len(page_candidates) >= 3:
                break
        if page_candidates:
            candidates.append(" ".join(page_candidates))

    # 3. Nombre del archivo como fallback garantizado
    filename_title = Path(str(pdf_path)).stem.replace("_", " ").replace("-", " ")
    candidates.append(filename_title)

    return candidates


SIMILARITY_THRESHOLD = 0.68  # Umbral de similitud para considerar duplicado


def _score_candidate(candidate: str, analyzed: dict) -> tuple:
    """Compara un candidato contra todos los analizados. Retorna (score, n, title)."""
    norm = _normalize_title(candidate)
    if not norm:
        return 0.0, None, None
    if norm in analyzed:
        n, title = analyzed[norm]
        return 1.0, n, title
    best_score = 0.0
    best_match = None
    for norm_key, (n, title) in analyzed.items():
        score = difflib.SequenceMatcher(None, norm, norm_key).ratio()
        if score > best_score:
            best_score = score
            best_match = (n, title)
    if best_match:
        return best_score, best_match[0], best_match[1]
    return 0.0, None, None


def find_duplicate(candidates: list, analyzed: dict) -> tuple:
    """
    Compara múltiples candidatos (metadata, contenido, nombre de archivo) contra los ya analizados.
    Retorna (es_duplicado, N°_existente, titulo_existente, similitud).
    """
    if not candidates or not analyzed:
        return False, None, None, 0.0

    best_score = 0.0
    best_n = None
    best_title = None

    for candidate in candidates:
        score, n, title = _score_candidate(candidate, analyzed)
        if score > best_score:
            best_score = score
            best_n = n
            best_title = title

    if best_score >= SIMILARITY_THRESHOLD and best_n is not None:
        return True, best_n, best_title, best_score

    return False, None, None, best_score

    return False, None, None, best_score


# ─────────────────────────── Auto-detección de inicio en Excel ───────────────────────────

def detect_start_number(excel_path: str) -> int:
    """Lee el Excel y retorna el siguiente N° a registrar (último N° existente + 1)."""
    _, last_num = load_analyzed_articles(excel_path)
    return last_num + 1


# ─────────────────────────── Reporte de procesamiento ───────────────────────────

REPORT_PATH = str(Path(__file__).parent / "reporte_procesamiento.txt")


def init_report(pdf_folder: str, start_number: int, total_pdfs: int):
    """Crea o sobreescribe el reporte con encabezado de sesión."""
    from datetime import datetime
    header = (
        "=" * 70 + "\n"
        "  REPORTE DE PROCESAMIENTO — REVISIÓN SISTEMÁTICA IaC\n"
        "=" * 70 + "\n"
        f"  Fecha:        {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"  Carpeta PDFs: {pdf_folder}\n"
        f"  Excel:        {EXCEL_PATH}\n"
        f"  Desde N°:     {start_number}  |  Total PDFs: {total_pdfs}\n"
        "=" * 70 + "\n\n"
        f"  {'N°':<6} {'Archivo PDF':<45} {'Título':<55} {'Decisión'}\n"
        f"  {'-'*6} {'-'*45} {'-'*55} {'-'*14}\n"
    )
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write(header)


def diagnostico_excel(excel_path: str, max_rows: int = 30):
    """Imprime el contenido bruto de las primeras filas del Excel para diagnóstico."""
    print(f"\n{'═'*62}")
    print(f"  DIAGNÓSTICO EXCEL — primeras {max_rows} filas")
    print(f"{'═'*62}")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[SHEET1_NAME]
        print(f"  {'Fila':<5} {'Col A (N°)':<15} {'Col B (Título)':<45} {'Col C (Autor)':<20}")
        print(f"  {'─'*4} {'─'*14} {'─'*44} {'─'*19}")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=3, values_only=True), 1):
            a, b, c = row[0], row[1], row[2]
            a_str = repr(a)[:14] if a is not None else "None"
            b_str = str(b)[:44] if b is not None else "None"
            c_str = str(c)[:19] if c is not None else "None"
            print(f"  {i:<5} {a_str:<15} {b_str:<45} {c_str:<20}")
        wb.close()
    except Exception as e:
        print(f"  Error al leer: {e}")
    print(f"{'═'*62}\n")


def append_to_report(art_num: int, pdf_name: str, tabla1: dict, error: str = ""):
    """Agrega una línea al reporte por cada artículo procesado."""
    if error:
        line = f"  {art_num:<6} {pdf_name:<45} {'— ERROR —':<55} {error[:40]}\n"
    else:
        titulo = tabla1.get("Artículo (Título)", "No especificado")[:53]
        decision = tabla1.get("Decisión", "N/A")
        line = f"  {art_num:<6} {pdf_name:<45} {titulo:<55} {decision}\n"
    with open(REPORT_PATH, "a", encoding="utf-8") as f:
        f.write(line)


def close_report(stats: dict):
    """Escribe el resumen final al reporte."""
    from datetime import datetime
    summary = (
        "\n" + "=" * 70 + "\n"
        "  RESUMEN FINAL\n"
        "=" * 70 + "\n"
        f"  Incluidos:       {stats['incluidos']}\n"
        f"  Rechazados F2:   {stats['rechazados_f2']}\n"
        f"  Ya analizados:   {stats.get('omitidos', 0)}  (omitidos por duplicado)\n"
        f"  Con error:       {len(stats['errores'])}\n"
    )
    if stats["errores"]:
        summary += "\n  Artículos con error:\n"
        for err in stats["errores"]:
            summary += f"    • {err}\n"
    summary += (
        f"\n  Finalizado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        + "=" * 70 + "\n"
    )
    with open(REPORT_PATH, "a", encoding="utf-8") as f:
        f.write(summary)


# ─────────────────────────── Main ───────────────────────────

def main():
    print("=" * 62)
    print("   PROCESADOR AUTOMÁTICO DE ARTÍCULOS ACADÉMICOS → EXCEL")
    print("   Motor de análisis: DeepSeek")
    print("=" * 62)
    print()

    # ── Validar configuración
    if DEEPSEEK_API_KEY.startswith("sk-xxx"):
        print("[ERROR] Debes ingresar tu API key de DeepSeek en la sección CONFIGURACIÓN del script.")
        sys.exit(1)

    pdf_folder = PDF_FOLDER.strip().strip('"')
    if not os.path.isdir(pdf_folder):
        print(f"[ERROR] Carpeta de PDFs no encontrada:\n  {pdf_folder}")
        print("  Verifica la variable PDF_FOLDER en la sección CONFIGURACIÓN.")
        sys.exit(1)

    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] No se encontró el Excel:\n  {EXCEL_PATH}")
        sys.exit(1)

    # Verificar que las hojas existan y mostrar diagnóstico si no
    try:
        _wb_check = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        _sheets = _wb_check.sheetnames
        _wb_check.close()
        missing_sheets = []
        if SHEET1_NAME not in _sheets:
            missing_sheets.append(SHEET1_NAME)
        if SHEET2_NAME not in _sheets:
            missing_sheets.append(SHEET2_NAME)
        if missing_sheets:
            print(f"[ERROR] No se encontraron estas hojas en el Excel:")
            for s in missing_sheets:
                print(f"  - '{s}'")
            print(f"\nHojas disponibles en el archivo:")
            for s in _sheets:
                print(f"  - '{s}'")
            print(f"\nCorrige SHEET1_NAME y SHEET2_NAME en la seccion CONFIGURACION del script.")
            sys.exit(1)
    except Exception as e:
        print(f"[ERROR] No se pudo abrir el Excel: {e}")
        sys.exit(1)

    api_key = DEEPSEEK_API_KEY

    # ── Diagnóstico: mostrar contenido bruto del Excel
    diagnostico_excel(EXCEL_PATH)

    # ── Leer artículos ya analizados en el Excel
    print("Leyendo artículos ya analizados en el Excel...")
    analyzed, last_num = load_analyzed_articles(EXCEL_PATH)
    start_number = last_num + 1

    print(f"\n{'─'*62}")
    print(f"  ARTÍCULOS YA ANALIZADOS EN EL EXCEL ({len(analyzed)} registros)")
    print(f"{'─'*62}")
    if analyzed:
        for norm_key, (n, title) in sorted(analyzed.items(), key=lambda x: x[1][0]):
            print(f"  N°{n:<5} {title[:70]}")
    else:
        print("  (ninguno — el Excel está vacío)")
    print(f"{'─'*62}")
    print(f"  Último N° con contenido: {last_num}")
    print(f"  Próximo N° a asignar:    {start_number}")

    # Advertir si hay saltos en la numeración (puede indicar filas corruptas)
    nums_found = sorted(n for (n, _) in analyzed.values())
    gaps = [nums_found[i] for i in range(1, len(nums_found)) if nums_found[i] - nums_found[i-1] > 1]
    if gaps:
        print(f"\n  ⚠  ATENCIÓN — Hay saltos en la numeración del Excel:")
        for g in gaps:
            prev = nums_found[nums_found.index(g) - 1]
            print(f"       Después de N°{prev} sigue N°{g}  (faltan N°{prev+1}–N°{g-1})")
        print(f"  Verifica el Excel: puede haber filas mal escritas o con título vacío.")
    print(f"{'─'*62}\n")

    # ── Listar PDFs ordenados por nombre
    pdf_files = sorted(Path(pdf_folder).glob("*.pdf"))
    total_pdfs = len(pdf_files)
    if total_pdfs == 0:
        print("[ERROR] No se encontraron archivos .pdf en la carpeta.")
        sys.exit(1)

    print(f"  PDFs en la carpeta ({total_pdfs} archivos):")
    for i, p in enumerate(pdf_files, 1):
        print(f"    {i:>3}. {p.name}")
    print()

    # ── Preguntar cuántos artículos NUEVOS analizar
    raw_limit = input(
        f"¿Cuántos artículos nuevos analizar?\n"
        f"  (recorre toda la carpeta y para cuando llegue a ese número)\n"
        f"  Ingresa un número o Enter para analizar todos ({total_pdfs}): "
    ).strip()
    if raw_limit.isdigit() and int(raw_limit) > 0:
        limit_new = int(raw_limit)
        print(f"  → Se procesarán hasta {limit_new} artículo(s) nuevos.\n")
    else:
        limit_new = 0  # 0 = sin límite
        print(f"  → Se procesarán todos los artículos nuevos que encuentre.\n")

    print(f"  Excel:   {EXCEL_PATH}")
    print(f"  Reporte: {REPORT_PATH}")
    print()

    # ── Cliente DeepSeek
    client = OpenAI(api_key=api_key, base_url=DEEPSEEK_BASE_URL)

    # ── Inicializar reporte
    init_report(pdf_folder, start_number, total_pdfs)

    # ── Estadísticas
    stats = {"incluidos": 0, "rechazados_f2": 0, "errores": [], "omitidos": 0}

    # ── Procesamiento por pares
    # CAMBIO: Verificar duplicados ANTES de asignar números, no después
    art_number = start_number
    pdf_index = 0
    new_processed = 0  # artículos nuevos efectivamente analizados
    pending_pair = []  # acumula PDFs nuevos para procesar en pares

    while pdf_index < total_pdfs or pending_pair:
        # ── Verificar si ya llegamos al límite de nuevos
        if limit_new > 0 and new_processed >= limit_new:
            print(f"\n  Límite de {limit_new} artículo(s) nuevos alcanzado. Deteniendo.")
            break

        # ── Recopilar PDFs nuevos hasta tener un par (2 elementos)
        while len(pending_pair) < 2 and pdf_index < total_pdfs:
            pdf_path = pdf_files[pdf_index]
            pdf_index += 1

            # Verificar si ya fue analizado
            print(f"\n  Verificando si '{pdf_path.name}' ya fue analizado...")
            title_candidates = extract_quick_title(pdf_path)
            for i, tc in enumerate(title_candidates):
                labels = ["metadata", "contenido", "archivo"]
                label = labels[i] if i < len(labels) else f"candidato {i+1}"
                print(f"    [{label}] {tc[:80]}")
            is_dup, dup_num, dup_title, similarity = find_duplicate(title_candidates, analyzed)

            if is_dup:
                # Es duplicado: omitir SIN asignar número
                pct = int(similarity * 100)
                print(f"  [OMITIDO] Ya analizado como N°{dup_num} ({pct}% similitud)")
                print(f"    Título en Excel: {dup_title[:80]}")
                append_to_report(
                    f"({pdf_path.name})", pdf_path.name,
                    {"Artículo (Título)": dup_title, "Decisión": f"YA ANALIZADO (N°{dup_num})"},
                    error=f"duplicado de N°{dup_num}"
                )
                stats["omitidos"] += 1
                continue  # Pasar al siguiente PDF SIN gastar número

            # Es nuevo: agregarlo al par pendiente CON número asignado
            pending_pair.append((art_number, pdf_path))
            art_number += 1

        # ── Si tenemos par completo (o es el último incompleto), procesarlo
        if pending_pair:
            pair = pending_pair[:2]  # máximo 2 elementos
            pending_pair = pending_pair[2:]  # guardar el resto

            pair_nums = [p[0] for p in pair]
            print()
            print("=" * 62)
            print(f"  PAR DE ARTÍCULOS: N°{pair_nums}")
            print("=" * 62)

            # Nueva conversación para este par
            messages = []
            n1, n2 = pair_nums[0], pair_nums[-1]

            print("Enviando Prompt Maestro a DeepSeek...")
            master = build_master_prompt(n1, n2)
            try:
                reply = send_to_deepseek(client, messages, master)
                print(f"  DeepSeek: {reply.strip()[:120]}")
            except Exception as e:
                print(f"  [ERROR] No se pudo iniciar conversación: {e}")
                for num, pdf_path in pair:
                    msg = f"fallo en prompt maestro — {str(e)[:60]}"
                    stats["errores"].append(f"N°{num}: {msg}")
                    append_to_report(num, pdf_path.name, {}, error=msg)
                continue

            # ── Procesar cada artículo del par
            for art_num, pdf_path in pair:
                tabla1, tabla2 = {}, {}

                try:
                    tabla1, tabla2 = process_article(
                        client, messages, art_num, pdf_path,
                        total=total_pdfs, current=new_processed + 1
                    )

                    # Guardar en Excel (xlwings escribe aunque el archivo esté abierto)
                    write_article_to_excel(EXCEL_PATH, art_num, tabla1, tabla2)

                    decision = tabla1.get("Decisión", "").upper()
                    if "INCLUIDO" in decision:
                        stats["incluidos"] += 1
                    elif "RECHAZADO" in decision:
                        stats["rechazados_f2"] += 1

                    new_processed += 1

                    # Agregar al índice en memoria para evitar duplicados en esta sesión
                    titulo_nuevo = tabla1.get("Artículo (Título)", "")
                    if titulo_nuevo:
                        analyzed[_normalize_title(titulo_nuevo)] = (art_num, titulo_nuevo)

                    append_to_report(art_num, pdf_path.name, tabla1)
                    print(f"  Decisión registrada: {tabla1.get('Decisión', 'N/A')}")

                except Exception as e:
                    err_msg = f"N°{art_num} ({pdf_path.name}): {str(e)[:80]}"
                    print(f"\n  [ERROR] {err_msg}")
                    stats["errores"].append(err_msg)
                    append_to_report(art_num, pdf_path.name, {}, error=str(e)[:60])

    # ── Cerrar reporte y mostrar resumen
    close_report(stats)

    print()
    print("=" * 62)
    print("  PROCESAMIENTO FINALIZADO")
    print("=" * 62)

    if new_processed == 0 and stats["omitidos"] > 0:
        print(f"  *** No se procesó ningún artículo nuevo ***")
        print(f"  Todos los PDFs revisados ya estaban analizados en el Excel.")
    else:
        print(f"  Nuevos procesados:   {new_processed}")
        print(f"  Incluidos:           {stats['incluidos']}")
        print(f"  Rechazados F2:       {stats['rechazados_f2']}")

    print(f"  Ya analizados:       {stats['omitidos']}  (omitidos por duplicado)")
    print(f"  Con error:           {len(stats['errores'])}")
    if stats["errores"]:
        print("\n  Artículos con error:")
        for err in stats["errores"]:
            print(f"    • {err}")
    print(f"\n  Reporte guardado en:\n  {REPORT_PATH}")
    print()


if __name__ == "__main__":
    main()
